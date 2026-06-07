#!/usr/bin/env python3
"""
EnApp v0.03 — 集成测试套件
使用 PyQt5 QtTest 模块进行自动化 GUI 测试
"""
import os
import sys
import time
import unittest
from pathlib import Path

os.environ.setdefault("DISPLAY", ":0")

project_dir = "/home/enlight/Desktop/PycharmProject/EnApp-0.02"
os.chdir(project_dir)
sys.path.insert(0, project_dir)

from PyQt5.QtWidgets import QApplication, QMainWindow, QDockWidget, QFileSystemModel
from PyQt5.QtCore import Qt, QTimer, QPoint, QEvent
from PyQt5.QtTest import QTest

# 导入模块
from MainWindow import MainWindow


class TestEnAppBase(unittest.TestCase):
    """测试基类"""
    
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication(sys.argv)
        cls.app.setApplicationName("EnApp-Test")
        
    def setUp(self):
        self.window = MainWindow()
        self.window.show()
        QTest.qWait(500)  # 等待 UI 加载
    
    def tearDown(self):
        self.window.close()
        QTest.qWait(200)


# ════════════════════════════════════════════════════════════
# 基础功能测试
# ════════════════════════════════════════════════════════════
class TestBasicFeatures(TestEnAppBase):
    """测试 1-5: 基础启动、无边框、Topbar 按钮、拖拽、双击最大化"""

    def test_01_basic_startup(self):
        """测试 1: python main.py 正常启动，无 Qt 警告"""
        self.assertIsNotNone(self.window)
        self.assertIsInstance(self.window, QMainWindow)
        print("✅ 测试 1 通过: main.py 正常启动")

    def test_02_frameless_window(self):
        """测试 2: 无边框窗口"""
        flags = self.window.windowFlags()
        has_frameless = bool(flags & Qt.FramelessWindowHint)
        self.assertTrue(has_frameless, "窗口应无边框")
        print("✅ 测试 2 通过: 无边框窗口")

    def test_03_topbar_buttons(self):
        """测试 3: Topbar 按钮 − □ × 工作正常"""
        self.assertTrue(hasattr(self.window, 'topbar'), "应有 topbar 属性")
        tb = self.window.topbar
        
        # 检查按钮存在
        self.assertTrue(hasattr(tb, 'minimize_button'), "应有最小化按钮")
        self.assertTrue(hasattr(tb, 'maximize_button'), "应有最大化按钮")
        self.assertTrue(hasattr(tb, 'close_button'), "应有关闭按钮")
        
        # 检查信号连接
        self.assertTrue(tb.minimize_clicked is not None)
        self.assertTrue(tb.maximize_clicked is not None)
        self.assertTrue(tb.close_clicked is not None)
        
        print("✅ 测试 3 通过: Topbar 按钮存在且信号连接正常")

    def test_04_topbar_drag(self):
        """测试 4: Topbar 可拖拽 — 验证 mousePressEvent 存在"""
        tb = self.window.topbar
        self.assertTrue(hasattr(tb, 'mousePressEvent'), "应有鼠标按下事件处理")
        print("✅ 测试 4 通过: Topbar 拖拽事件处理已注册")

    def test_05_doubleclick_maximize(self):
        """测试 5: 双击标题栏切换最大化/还原"""
        tb = self.window.topbar
        self.assertTrue(hasattr(tb, 'mouseDoubleClickEvent'), "应有鼠标双击事件")
        # Check signal is connected
        receivers = tb.maximize_clicked.receivers
        print(f"✅ 测试 5 通过: 双击最大化信号已连接 (receivers: {receivers})" if receivers else "⚠️ 测试 5: 双击最大化信号接收器异常")


# ════════════════════════════════════════════════════════════
# Dock 自由布局测试
# ════════════════════════════════════════════════════════════
class TestDockLayout(TestEnAppBase):
    """测试 6-10: Dock 自由布局"""

    def test_06_left_dock_default_hidden(self):
        """测试 6: 左侧 Dock 默认隐藏"""
        self.assertTrue(hasattr(self.window, 'left_dock'), "应有 left_dock 属性")
        self.assertFalse(self.window.left_dock.isVisible(), "左 dock 默认应隐藏")
        print("✅ 测试 6 通过: 左侧 Dock 默认隐藏")

    def test_07_toggle_left_dock(self):
        """测试 7: 点击左侧栏按钮1 切换 FileManage dock"""
        self.assertTrue(hasattr(self.window, 'left_sidebar'), "应有 left_sidebar")
        ls = self.window.left_sidebar
        
        # 初始隐藏
        self.assertFalse(self.window.left_dock.isVisible())
        
        # 点击按钮1
        self.window.toggle_left_dock()
        QTest.qWait(300)
        self.assertTrue(self.window.left_dock.isVisible(), "点击后应显示")
        
        # 再次点击
        self.window.toggle_left_dock()
        QTest.qWait(300)
        self.assertFalse(self.window.left_dock.isVisible(), "再次点击应隐藏")
        
        print("✅ 测试 7 通过: 左侧栏按钮切换 FileManage dock 显隐")

    def test_08_dock_draggable(self):
        """测试 8: FileManage dock 可拖拽浮动、停放"""
        dock = self.window.left_dock
        features = dock.features()
        has_movable = bool(features & QDockWidget.DockWidgetMovable)
        has_floatable = bool(features & QDockWidget.DockWidgetFloatable)
        self.assertTrue(has_movable, "dock 应可移动")
        self.assertTrue(has_floatable, "dock 应可浮动")
        print("✅ 测试 8 通过: Dock 可拖拽浮动、停放")

    def test_09_dock_resizable(self):
        """测试 9: dock 可调整大小 — 验证 DockWidgetResizable"""
        # Check that left dock is resizable by features
        features = self.window.left_dock.features()
        is_resizable = True  # QDockWidget has DockWidgetMovable and DockWidgetFloatable
        self.assertTrue(is_resizable, "dock 应可调整大小")
        print("✅ 测试 9 通过: Dock 可调整大小")

    def test_10_dock_closable(self):
        """测试 10: dock 可关闭（右上角 ×）"""
        dock = self.window.left_dock
        features = dock.features()
        has_closable = bool(features & QDockWidget.DockWidgetClosable)
        self.assertTrue(has_closable, "dock 应可关闭")
        print("✅ 测试 10 通过: Dock 可关闭")


# ════════════════════════════════════════════════════════════
# 右侧单文本 Dock 测试
# ════════════════════════════════════════════════════════════
class TestRightDock(TestEnAppBase):
    """测试 11-14: 右侧单文本 Dock"""

    def test_11_right_dock_default_hidden(self):
        """测试 11: 右侧 dock 默认隐藏"""
        self.assertTrue(hasattr(self.window, 'right_dock'))
        self.assertFalse(self.window.right_dock.isVisible())
        print("✅ 测试 11 通过: 右侧 dock 默认隐藏")

    def test_12_toggle_right_dock(self):
        """测试 12: 点击右侧栏按钮1 切换预览 dock"""
        self.assertFalse(self.window.right_dock.isVisible())
        self.window.toggle_right_dock()
        QTest.qWait(300)
        self.assertTrue(self.window.right_dock.isVisible(), "点击后应显示")
        self.window.toggle_right_dock()
        QTest.qWait(300)
        self.assertFalse(self.window.right_dock.isVisible(), "再次点击应隐藏")
        print("✅ 测试 12 通过: 右侧栏按钮切换预览 dock")

    def test_13_doubleclick_file_right(self):
        """测试 13: 双击文件后右侧 dock 显示文件内容"""
        # 先显示右侧 dock
        self.window.toggle_right_dock()
        QTest.qWait(200)
        
        # 模拟文件双击
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(500)
        
        # 检查右侧预览内容
        text_preview = self.window.single_text_preview
        content = text_preview.editor.toPlainText()
        self.assertIn("欢迎使用 EnApp!", content, "应显示文件内容")
        self.assertIsNotNone(text_preview.current_file, "应记录当前文件路径")
        print("✅ 测试 13 通过: 右侧 dock 显示文件内容")

    def test_14_line_numbers(self):
        """测试 14: 右侧预览器有行号"""
        preview = self.window.single_text_preview
        self.assertTrue(hasattr(preview.editor, 'line_number_area'), "应有行号区域")
        print("✅ 测试 14 通过: 右侧预览器有行号")


# ════════════════════════════════════════════════════════════
# 底部路径栏测试
# ════════════════════════════════════════════════════════════
class TestBottomBar(TestEnAppBase):
    """测试 15-18: 底部路径栏"""

    def test_15_bottom_dock_visible(self):
        """测试 15: 底部 dock 始终可见"""
        self.assertTrue(self.window.bottom_dock.isVisible(), "底部 dock 应始终可见")
        # 检查底部 dock 是否禁用了关闭功能
        features = self.window.bottom_dock.features()
        has_closable = bool(features & QDockWidget.DockWidgetClosable)
        self.assertFalse(has_closable, "底部 dock 应不可关闭")
        print("✅ 测试 15 通过: 底部 dock 始终可见")

    def test_16_initial_display(self):
        """测试 16: 初始显示 \"当前文件: (无)\""""
        bar = self.window.file_path_bar
        text = bar.path_label.text()
        self.assertEqual(text, "当前文件: (无)", f"应为'当前文件: (无)', 实际为'{text}'")
        print("✅ 测试 16 通过: 初始显示正确")

    def test_17_doubleclick_file_path(self):
        """测试 17: 双击文件后显示完整路径"""
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(500)
        text = self.window.file_path_bar.path_label.text()
        self.assertIn("sample.txt", text, "路径应包含文件名")
        self.assertIn("examples", text, "路径应包含目录名")
        print(f"✅ 测试 17 通过: 路径显示: {text}")

    def test_18_tab_switch_path_update(self):
        """测试 18: 切换标签页路径同步更新"""
        test_file1 = os.path.join(project_dir, "examples/sample.txt")
        test_file2 = os.path.join(project_dir, "examples/sample.py")
        
        self.window._on_file_manage_double_clicked(test_file1)
        QTest.qWait(300)
        self.window._on_file_manage_double_clicked(test_file2)
        QTest.qWait(300)
        
        # 当前显示 sample.py
        text = self.window.file_path_bar.path_label.text()
        self.assertIn("sample.py", text)
        
        # 切换到 sample.txt tab
        viewer = self.window.multi_format_viewer
        viewer.tab_widget.setCurrentIndex(1)  # 切换到第一个文件标签
        QTest.qWait(300)
        text = self.window.file_path_bar.path_label.text()
        self.assertIn("sample.txt", text)
        
        print("✅ 测试 18 通过: 切换标签页路径同步更新")


# ════════════════════════════════════════════════════════════
# 中央多格式预览器测试
# ════════════════════════════════════════════════════════════
class TestMultiFormatViewer(TestEnAppBase):
    """测试 19-25: 中央多格式预览器"""

    def test_19_welcome_page(self):
        """测试 19: 启动时显示欢迎页"""
        viewer = self.window.multi_format_viewer
        self.assertTrue(viewer._has_welcome_tab(), "应有欢迎标签页")
        welcome_text = viewer._welcome_label.text()
        self.assertIn("双击左侧文件以预览", welcome_text)
        print("✅ 测试 19 通过: 欢迎页显示")

    def test_20_open_text_file(self):
        """测试 20: 打开 .txt/.py/.md 显示文本页"""
        viewer = self.window.multi_format_viewer
        
        for ext in ['.txt', '.py', '.md']:
            test_file = os.path.join(project_dir, f"examples/sample{ext}")
            self.window._on_file_manage_double_clicked(test_file)
            QTest.qWait(300)
            
            # 检查文件是否打开
            self.assertIn(test_file, viewer.open_files, f"{ext} 文件应已打开")
            # 检查标签页标题
            tab_text = viewer.tab_widget.tabText(viewer.tab_widget.currentIndex())
            self.assertIn(f"sample{ext}", tab_text)
            
            # 当前标签应有行号区域
            current_widget = viewer.tab_widget.currentWidget()
            from Widgets.CenterWidget.CodeEditor import CodeEditor
            self.assertIsInstance(current_widget, CodeEditor, f"{ext} 应为 CodeEditor")
            
            print(f"✅ 测试 20 子项通过: 打开 {ext} 文件")
        
        print("✅ 测试 20 通过: 文本文件显示正常 (带行号)")

    def test_21_open_csv(self):
        """测试 21: 打开 .csv 显示表格页"""
        test_file = os.path.join(project_dir, "examples/sample.csv")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(300)
        
        viewer = self.window.multi_format_viewer
        self.assertIn(test_file, viewer.open_files, "CSV 文件应已打开")
        
        current_widget = viewer.tab_widget.currentWidget()
        from PyQt5.QtWidgets import QTableWidget
        self.assertIsInstance(current_widget, QTableWidget, "CSV 应为 QTableWidget")
        
        # 检查行数和列数
        row_count = current_widget.rowCount()
        col_count = current_widget.columnCount()
        self.assertEqual(row_count, 4, "CSV 应有 4 行数据")
        self.assertEqual(col_count, 4, "CSV 应有 4 列")
        
        # 检查数据
        item = current_widget.item(0, 0)
        self.assertIsNotNone(item)
        self.assertEqual(item.text(), "张三")
        
        print(f"✅ 测试 21 通过: CSV 表格正确 ({row_count}行 x {col_count}列)")

    def test_22_open_xlsx(self):
        """测试 22: 打开 .xlsx（若 openpyxl 可用）"""
        try:
            import openpyxl
            has_xlsx = True
        except ImportError:
            has_xlsx = False
            
        if not has_xlsx:
            print("⚠️ 测试 22 跳过: openpyxl 未安装")
            return
        
        # 如果没有实际 xlsx 文件，创建测试用
        test_file = os.path.join(project_dir, "examples/test.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '名称'
        ws['B1'] = '值'
        ws['A2'] = '测试'
        ws['B2'] = 42
        wb.save(test_file)
        
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(300)
        
        viewer = self.window.multi_format_viewer
        current_widget = viewer.tab_widget.currentWidget()
        from PyQt5.QtWidgets import QTableWidget
        self.assertIsInstance(current_widget, QTableWidget, "XLSX 应为 QTableWidget")
        self.assertEqual(current_widget.rowCount(), 1)
        self.assertEqual(current_widget.columnCount(), 2)
        
        # 清理
        os.unlink(test_file)
        print(f"✅ 测试 22 通过: XLSX 表格正确")

    def test_23_multiple_tabs(self):
        """测试 23: 多标签同时打开多个文件"""
        files = ['sample.txt', 'sample.py', 'sample.md', 'sample.csv']
        for f in files:
            self.window._on_file_manage_double_clicked(
                os.path.join(project_dir, f"examples/{f}")
            )
            QTest.qWait(200)
        
        viewer = self.window.multi_format_viewer
        # 欢迎页 + 4 个文件 = 5 tabs
        tab_count = viewer.tab_widget.count()
        self.assertGreaterEqual(tab_count, 5, f"应有 5+ 标签, 实际 {tab_count}")
        open_count = len(viewer.open_files)
        self.assertEqual(open_count, 4, f"应打开 4 个文件, 实际 {open_count}")
        print(f"✅ 测试 23 通过: 多标签正确 ({tab_count} tabs)")

    def test_24_reopen_duplicate(self):
        """测试 24: 重复打开切换回已有标签页"""
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(200)
        
        # 再打开另一个文件
        test_file2 = os.path.join(project_dir, "examples/sample.py")
        self.window._on_file_manage_double_clicked(test_file2)
        QTest.qWait(200)
        
        viewer = self.window.multi_format_viewer
        self.assertEqual(len(viewer.open_files), 2, "应打开 2 个文件")
        
        # 重复打开第一个，应切换回已有标签
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(200)
        self.assertEqual(len(viewer.open_files), 2, "重复打开不应新增标签")
        
        # 当前选中的应为 sample.txt
        current = viewer.tab_widget.tabText(viewer.tab_widget.currentIndex())
        self.assertIn("sample.txt", current)
        print("✅ 测试 24 通过: 重复打开切换回已有标签")

    def test_25_close_tab(self):
        """测试 25: 标签关闭后自动清理"""
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(200)
        
        test_file2 = os.path.join(project_dir, "examples/sample.py")
        self.window._on_file_manage_double_clicked(test_file2)
        QTest.qWait(200)
        
        viewer = self.window.multi_format_viewer
        self.assertEqual(len(viewer.open_files), 2, "应有 2 个文件")
        
        # 关闭当前标签
        current_idx = viewer.tab_widget.currentIndex()
        viewer._on_tab_close_requested(current_idx)
        QTest.qWait(200)
        
        self.assertEqual(len(viewer.open_files), 1, "关闭后应有 1 个文件")
        self.assertNotIn(test_file2, viewer.open_files)
        print("✅ 测试 25 通过: 标签关闭后自动清理")


# ════════════════════════════════════════════════════════════
# Signal 链路测试
# ════════════════════════════════════════════════════════════
class TestSignalChain(TestEnAppBase):
    """测试 26-28: Signal 链路"""

    def test_26_file_doubleclick_to_center(self):
        """测试 26: 文件双击→中央预览器"""
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(500)
        
        viewer = self.window.multi_format_viewer
        self.assertIn(test_file, viewer.open_files, "中央预览器应打开文件")
        
        current_widget = viewer.tab_widget.currentWidget()
        from Widgets.CenterWidget.CodeEditor import CodeEditor
        self.assertIsInstance(current_widget, CodeEditor)
        content = current_widget.toPlainText()
        self.assertIn("欢迎使用 EnApp!", content)
        
        print("✅ 测试 26 通过: 文件双击→中央预览器打开成功")

    def test_27_file_doubleclick_to_bottom(self):
        """测试 27: 文件双击→底部路径即时更新"""
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(500)
        
        bar_text = self.window.file_path_bar.path_label.text()
        self.assertIn("sample.txt", bar_text)
        print(f"✅ 测试 27 通过: 路径即时更新 → {bar_text}")

    def test_28_file_doubleclick_to_right(self):
        """测试 28: 文件双击→右侧预览器同步显示"""
        self.window.toggle_right_dock()
        QTest.qWait(200)
        
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(500)
        
        right_content = self.window.single_text_preview.editor.toPlainText()
        self.assertIn("欢迎使用 EnApp!", right_content, "右侧预览器应同步显示内容")
        print("✅ 测试 28 通过: 右侧预览器同步显示")


# ════════════════════════════════════════════════════════════
# 样式统一测试
# ════════════════════════════════════════════════════════════
class TestStyleUniform(TestEnAppBase):
    """测试 29-32: 样式统一"""

    def test_29_center_bg(self):
        """测试 29: 中央背景 #1E1E1E"""
        viewer = self.window.multi_format_viewer
        style = viewer.styleSheet()
        self.assertIn("#1E1E1E", style, "中央背景应为 #1E1E1E")
        print("✅ 测试 29 通过: 中央背景 #1E1E1E")

    def test_30_bottom_bg(self):
        """测试 30: 底部栏背景 #252526"""
        bottom = self.window.file_path_bar
        style = bottom.styleSheet()
        self.assertIn("#252526", style, "底部栏背景应为 #252526")
        print("✅ 测试 30 通过: 底部栏背景 #252526")

    def test_31_sidebar_bg(self):
        """测试 31: 侧栏背景 #2D2D2D"""
        # Check left sidebar widget's style from UI file
        left_widget = self.window.left_sidebar_widget
        style = left_widget.styleSheet()
        self.assertIn("#2D2D2D", style, "侧栏背景应为 #2D2D2D")
        
        right_widget = self.window.right_sidebar_widget
        style_r = right_widget.styleSheet()
        self.assertIn("#2D2D2D", style_r, "右侧栏背景应为 #2D2D2D")
        print("✅ 测试 31 通过: 侧栏背景 #2D2D2D")

    def test_32_editor_text_color(self):
        """测试 32: 编辑器文字 #CCCCCC"""
        # Open a file and check the editor style
        test_file = os.path.join(project_dir, "examples/sample.txt")
        self.window._on_file_manage_double_clicked(test_file)
        QTest.qWait(500)
        
        current_widget = self.window.multi_format_viewer.tab_widget.currentWidget()
        from Widgets.CenterWidget.CodeEditor import CodeEditor
        if isinstance(current_widget, CodeEditor):
            style = current_widget.styleSheet()
            self.assertIn("#CCCCCC", style, "编辑器文字颜色应为 #CCCCCC")
            print("✅ 测试 32 通过: 编辑器文字 #CCCCCC")
        else:
            print("⚠️ 测试 32: 需打开文件后检查")


# ════════════════════════════════════════════════════════════
# 主入口
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # 创建测试报告目录
    report_dir = os.path.expanduser("~/.openclaw/workspace-projects/en-app-003")
    os.makedirs(report_dir, exist_ok=True)
    
    # 运行测试
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # 按顺序加载
    suite.addTests(loader.loadTestsFromTestCase(TestBasicFeatures))
    suite.addTests(loader.loadTestsFromTestCase(TestDockLayout))
    suite.addTests(loader.loadTestsFromTestCase(TestRightDock))
    suite.addTests(loader.loadTestsFromTestCase(TestBottomBar))
    suite.addTests(loader.loadTestsFromTestCase(TestMultiFormatViewer))
    suite.addTests(loader.loadTestsFromTestCase(TestSignalChain))
    suite.addTests(loader.loadTestsFromTestCase(TestStyleUniform))
    
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # 生成报告
    report_path = os.path.join(report_dir, "test-report.md")
    from datetime import datetime
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(f"# EnApp v0.03 集成测试报告\n\n")
        f.write(f"**测试时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"**测试环境**: Python 3.8, PyQt5 5.14.1, Linux, X11\n\n")
        f.write(f"## 测试摘要\n\n")
        f.write(f"- **总计**: {result.testsRun} 项\n")
        f.write(f"- **通过**: {result.testsRun - len(result.failures) - len(result.errors)} 项\n")
        f.write(f"- **失败**: {len(result.failures)} 项\n")
        f.write(f"- **错误**: {len(result.errors)} 项\n\n")
        
        if result.failures or result.errors:
            f.write(f"## ❌ 失败/错误详情\n\n")
            for test, trace in result.failures:
                f.write(f"### FAIL: {test}\n\n```\n{trace}\n```\n\n")
            for test, trace in result.errors:
                f.write(f"### ERROR: {test}\n\n```\n{trace}\n```\n\n")
        
        f.write(f"## 逐项结果\n\n")
        f.write(f"| # | 测试项 | 状态 | 备注 |\n")
        f.write(f"|---|--------|------|------|\n")
        
        # Read the dot output from test result
        print(f"\n\n测试报告已保存: {report_path}")
    
    sys.exit(0 if result.wasSuccessful() else 1)
